# --------------------------------------------------------------------------------------------------
# THIS FILE COLLECTS DATA FROM THE EXCEL FILES THEN INSERTS THE INFORMATION INTO A SINGLE EXCEL FILE
# --------------------------------------------------------------------------------------------------
from openpyxl import load_workbook, Workbook
from datetime import datetime, date
import shutil
import os 


# PREREQUISITE FOR SCRIPT
# 1. NEED THE PATH OF THE FOLDER


# -------------------------------------------------------------------------------------------------
# this is the path of the folder: must be updated to the folder with the Excel files
folder_path = "C:\\Users\\BTruong\\OneDrive - TRC\\Documents\\Master File Python\\00000000_template"
# --------------------------------------------------------------------------------------------------


def return_list_of_paths(folder_path:str):
    """
    Purpose:
        Returns a list of paths to all Excel files within the folder and its subfolders.

    Parameters:
        folder_path (str): The path of the folder that where searching for Excel files will start.

    Returns:
        list: A list of file paths to all the Excel files found within the folder and its subfolders.
    """
    path_list = []
    return_path_list = []
    folder = os.listdir(folder_path)

    # access the folder
    for file in folder:
        # create the path for the folder and enter the folder
        excel_folder_path = folder_path + "\\" + file
        excel_folder_path_original = excel_folder_path
        if os.path.isdir(excel_folder_path):
            excel_folder = os.listdir(excel_folder_path)

            # in each folder check for excel files
            for excel_file in excel_folder:
                if excel_file[-4:] == "xlsx":
                    excel_folder_path += "\\" + excel_file
                    path_list.append(excel_folder_path)
                excel_folder_path = excel_folder_path_original


    # replace '\' with '\\'
    for path in path_list:
        path = path.replace("\\", "\\\\")
        return_path_list.append(path)

    return return_path_list


def create_excel_file(path:str):
    """
        Purpose:
            Creates a new Excel file with a predefined header structure.

        Parameters:
            path (str): This is the file path where the Excel file will be created.

        Returns:
            None: This function does not return anything, but it creates the Excel file at the specified 'path'.

        Example:
            create_excel_file('C:/Users/username/Documents/my_excel_file.xlsx')
    """
    workbook = Workbook()
    worksheet = workbook['Sheet']

    # naming the header columns
    worksheet['A1'] = 'Asset ID'
    worksheet['B1'] = 'Tab Name'
    worksheet['C1'] = 'Table Name'
    worksheet['D1'] = 'Parent Table ID'
    worksheet['E1'] = 'Child Table ID' 
    worksheet['F1'] = 'Data Validation' 
    worksheet['G1'] = 'Enclosure 1'
    worksheet['H1'] = 'Enclosure 2'
    worksheet['I1'] = 'Photo A Interior Box Surfaces'
    worksheet['J1'] = 'Photo B Doors'
    worksheet['K1'] = 'Photo C Outside of the Box'
    worksheet['L1'] = 'Photo D Clear Photos of all Terminal Blocks'
    worksheet['M1'] = 'Photo E Power'
    worksheet['N1'] = 'Photo F Telecom'
    worksheet['O1'] = 'Length'
    worksheet['P1'] = 'L Unit'
    worksheet['Q1'] = 'Width'
    worksheet['R1'] = 'W Unit'
    worksheet['S1'] = 'Qty Conduits'
    worksheet['T1'] = 'Qty Inputs'
    worksheet['U1'] = 'Types of Inputs'
    worksheet['V1'] = 'Types of Conduits'
    worksheet['W1'] = 'Counter'

    worksheet['AA1'] = 'Asset ID'
    worksheet['AB1'] = 'Tab Name'
    worksheet['AC1'] = 'Table Name'
    worksheet['AD1'] = 'Parent Table ID'
    worksheet['AE1'] = 'Child Table ID' 
    worksheet['AF1'] = 'Data Validation' 
    worksheet['AG1'] = 'Power Type' 
    worksheet['AH1'] = 'Type' 
    worksheet['AI1'] = 'Voltage Rating' 
    worksheet['AJ1'] = 'Max Capacity of Device' 
    worksheet['AK1'] = 'Qty In Use' 
    worksheet['AL1'] = 'Qty Available' 
    worksheet['AM1'] = 'Counter'

    worksheet['AO1'] = 'Asset ID'
    worksheet['AP1'] = 'Tab Name'
    worksheet['AQ1'] = 'Table Name'
    worksheet['AR1'] = 'Parent Table ID'
    worksheet['AS1'] = 'Child Table ID' 
    worksheet['AT1'] = 'Data Validation' 
    worksheet['AU1'] = 'Breaker Type'
    worksheet['AV1'] = 'Device Type'
    worksheet['AW1'] = 'Rating'
    worksheet['AX1'] = 'Brand/Model'
    worksheet['AY1'] = 'Counter'

    worksheet['BA1'] = 'Asset ID'
    worksheet['BB1'] = 'Tab Name'
    worksheet['BC1'] = 'Table Name'
    worksheet['BD1'] = 'Parent Table ID'
    worksheet['BE1'] = 'Child Table ID' 
    worksheet['BF1'] = 'Data Validation' 
    worksheet['BG1'] = 'Instrumentation Type' 
    worksheet['BH1'] = 'Make' 
    worksheet['BI1'] = 'Model' 
    worksheet['BJ1'] = 'Serial No' 
    worksheet['BK1'] = 'Reference to the glossary list' 
    worksheet['BL1'] = 'Counter'

    worksheet['CA1'] = 'Asset ID'
    worksheet['CB1'] = 'Tab Name'
    worksheet['CC1'] = 'Table Name'
    worksheet['CD1'] = 'Parent Table ID'
    worksheet['CE1'] = 'Child Table ID' 
    worksheet['CF1'] = 'Data Validation' 
    worksheet['CG1'] = 'Type PLC.RTU' 
    worksheet['CH1'] = 'Asset Serial Number' 
    worksheet['CI1'] = 'Asset Tag' 
    worksheet['CJ1'] = 'RTU Use' 
    worksheet['CK1'] = 'Counter'

    worksheet['CM1'] = 'Asset ID'
    worksheet['CN1'] = 'Tab Name'
    worksheet['CO1'] = 'Table Name'
    worksheet['CP1'] = 'Parent Table ID'
    worksheet['CQ1'] = 'Child Table ID' 
    worksheet['CR1'] = 'Data Validation' 
    worksheet['CS1'] = 'Port Type' 
    worksheet['CT1'] = 'Protocol' 
    worksheet['CU1'] = 'Status' 
    worksheet['CV1'] = 'Comment Description' 
    worksheet['CW1'] = 'Counter'

    worksheet['DA1'] = 'Asset ID'
    worksheet['DB1'] = 'Tab Name'
    worksheet['DC1'] = 'Table Name'
    worksheet['DD1'] = 'Parent Table ID'
    worksheet['DE1'] = 'Child Table ID' 
    worksheet['DF1'] = 'Data Validation' 
    worksheet['DG1'] = 'IO Classification 1' 
    worksheet['DH1'] = 'IO Classification 2' 
    worksheet['DI1'] = 'Qty of IO Ports' 
    worksheet['DJ1'] = 'Qty of IO Ports in use' 
    worksheet['DK1'] = 'Label'
    worksheet['DL1'] = 'Counter'

    worksheet['EA1'] = 'Asset ID'
    worksheet['EB1'] = 'Tab Name'
    worksheet['EC1'] = 'Table Name'
    worksheet['ED1'] = 'Parent Table ID'
    worksheet['EE1'] = 'Child Table ID' 
    worksheet['EF1'] = 'Data Validation'
    worksheet['EG1'] = 'Comms Type 1'
    worksheet['EH1'] = 'Comms Type 2'
    worksheet['EI1'] = 'Device Type 1'
    worksheet['EJ1'] = 'Device Type 2'
    worksheet['EK1'] = 'Model Number'
    worksheet['EL1'] = 'Frequency'
    worksheet['EM1'] = 'Wireless Strength'
    worksheet['EN1'] = 'Counter'

    worksheet['EP1'] = 'Asset ID'
    worksheet['EQ1'] = 'Tab Name'
    worksheet['ER1'] = 'Table Name'
    worksheet['ES1'] = 'Parent Table ID'
    worksheet['ET1'] = 'Child Table ID' 
    worksheet['EU1'] = 'Data Validation'
    worksheet['EV1'] = 'Port Type'
    worksheet['EW1'] = 'Protocol'
    worksheet['EX1'] = 'Status'
    worksheet['EY1'] = 'Comment Description'
    worksheet['EZ1'] = 'Counter'

    worksheet['FB1'] = 'Asset ID'
    worksheet['FC1'] = 'Counter'
    worksheet['FD1'] = 'File Name'
    worksheet['FE1'] = 'Last Date Modified'
    worksheet['FF1'] = 'Facility Name'
    worksheet['FG1'] = 'Facility Name Confirmation'
    worksheet['FH1'] = 'Address'
    worksheet['FI1'] = 'Address Confirmation'
    worksheet['FJ1'] = 'Latitude'
    worksheet['FK1'] = 'Latitude Confirmation'
    worksheet['FL1'] = 'Longitude'
    worksheet['FM1'] = 'Longitude Confirmation'
    worksheet['FN1'] = 'Input from System'
    worksheet['FO1'] = 'Input from System Confirmation'
    worksheet['FP1'] = 'Output to System'
    worksheet['FQ1'] = 'Output to System Confirmation'
    worksheet['FR1'] = 'Closest Pole'
    worksheet['FS1'] = 'Closest Pole Confirmation'
    worksheet['FT1'] = 'Expected Wireframe Type'
    worksheet['FU1'] = 'Expected Wireframe Type Confirmation'
    worksheet['FV1'] = 'Inlet CV'
    worksheet['FW1'] = 'Inlet CV Confirmation'
    worksheet['FX1'] = 'Outlet CV'
    worksheet['FY1'] = 'Outlet CV Confirmation'
    worksheet['FZ1'] = 'Cellular Strength'
    worksheet['GA1'] = 'Cellular Strength Confirmation'
    worksheet['GB1'] = 'Version'

    worksheet['GD1'] = '1.Enclosure Parent Index'
    worksheet['GE1'] = '2.Power Parent Index'
    worksheet['GF1'] = '2.Power Comms Index'
    worksheet['GG1'] = '3.Instrumentation Parent Index'
    worksheet['GH1'] = '4.PLC Parent Index'
    worksheet['GI1'] = '4.PLC Comms Index'
    worksheet['GJ1'] = '4.PLC IO Index'
    worksheet['GK1'] = '5.Comms Parent Index'
    worksheet['GL1'] = '5.Comms Parent Index'
    worksheet['GM1'] = '0.General Parent Index'
    worksheet['GN1'] = 'Counter Index'
    worksheet['GD2'] = '2'
    worksheet['GE2'] = '2'
    worksheet['GF2'] = '2'
    worksheet['GG2'] = '2'
    worksheet['GH2'] = '2'
    worksheet['GI2'] = '2'
    worksheet['GJ2'] = '2'
    worksheet['GK2'] = '2'
    worksheet['GL2'] = '2'
    worksheet['GM2'] = '2'
    worksheet['GN2'] = '1'

    workbook.save(path)


def retrieve_asset_ID(path:str):
    """
        Purpose:
            Retrieves the asset ID from the specified Excel file
            
        Parameters:
            path (str): The path of the Excel file to retrieve the asset ID

        Returns:
            int: The asset ID extracted from the Excel file. Returns 0 if the asset ID is null.
    """
    excel_file_wb = load_workbook(path)
    general_ws = excel_file_wb['0.General']
    
    asset_ID = general_ws['C4'].value
    if asset_ID is None:
        asset_ID = 0

    return asset_ID


def general_extraction(path:str):
    """
        Purpose: 
            Extracts data from an Excel file in the Enclosure sheet and writes it to a consolidated Excel file.

        Parameters: 
            path (str): This is the file path of the Excel file.

        Returns:
            None: This function does not return anything, but it writes to the consolidated Excel file. 
    """
    excel_file_wb = load_workbook(path)
    general_ws = excel_file_wb['0.General']

    consolidated_wb = load_workbook('Wireframe_Consolidated_Data.xlsx')
    consolidated_ws = consolidated_wb['Sheet']


    # asset ID
    asset_ID = retrieve_asset_ID(path)


    # file name
    file_name = os.path.basename(path)
    file_name = file_name[:-5]


    # row
    row = int(consolidated_ws['GM2'].value)


    # last date modified
    modification_time = os.path.getmtime(path)
    last_modified_date = datetime.fromtimestamp(modification_time)


    # recieving counter
    counter = consolidated_ws['GN2'].value


    # checking version
    version = 0
    if general_ws['I3'].value == 'Inlet CV':
        version = 1
    else:
        version = 2


    # inserting data
    if version == 1:
        consolidated_ws[f"FB{row}"].value = asset_ID
        consolidated_ws[f"FC{row}"].value = counter
        consolidated_ws[f"FD{row}"].value = file_name
        consolidated_ws[f"FE{row}"].value = last_modified_date
        consolidated_ws[f"FF{row}"].value = general_ws['B4'].value
        consolidated_ws[f"FG{row}"].value = general_ws['B5'].value
        consolidated_ws[f"FH{row}"].value = general_ws['D4'].value
        consolidated_ws[f"FI{row}"].value = general_ws['D5'].value
        consolidated_ws[f"FJ{row}"].value = general_ws['E4'].value
        consolidated_ws[f"FK{row}"].value = general_ws['E5'].value
        consolidated_ws[f"FL{row}"].value = general_ws['F4'].value
        consolidated_ws[f"FM{row}"].value = general_ws['F5'].value
        consolidated_ws[f"FN{row}"].value = general_ws['G4'].value
        consolidated_ws[f"FO{row}"].value = general_ws['G5'].value
        consolidated_ws[f"FP{row}"].value = general_ws['H4'].value
        consolidated_ws[f"FQ{row}"].value = general_ws['H5'].value
        consolidated_ws[f"FV{row}"].value = general_ws['I4'].value
        consolidated_ws[f"FW{row}"].value = general_ws['I5'].value
        consolidated_ws[f"FX{row}"].value = general_ws['J4'].value
        consolidated_ws[f"FY{row}"].value = general_ws['J5'].value
        consolidated_ws[f"FZ{row}"].value = general_ws['K4'].value
        consolidated_ws[f"GA{row}"].value = general_ws['K5'].value
        consolidated_ws[f"GB{row}"].value = version
    if version == 2:
        consolidated_ws[f"FB{row}"].value = asset_ID
        consolidated_ws[f"FC{row}"].value = counter
        consolidated_ws[f"FD{row}"].value = file_name
        consolidated_ws[f"FE{row}"].value = last_modified_date
        consolidated_ws[f"FF{row}"].value = general_ws['B4'].value
        consolidated_ws[f"FG{row}"].value = general_ws['B5'].value
        consolidated_ws[f"FH{row}"].value = general_ws['D4'].value
        consolidated_ws[f"FI{row}"].value = general_ws['D5'].value
        consolidated_ws[f"FJ{row}"].value = general_ws['E4'].value
        consolidated_ws[f"FK{row}"].value = general_ws['E5'].value
        consolidated_ws[f"FL{row}"].value = general_ws['F4'].value
        consolidated_ws[f"FM{row}"].value = general_ws['F5'].value
        consolidated_ws[f"FN{row}"].value = general_ws['G4'].value
        consolidated_ws[f"FO{row}"].value = general_ws['G5'].value
        consolidated_ws[f"FP{row}"].value = general_ws['H4'].value
        consolidated_ws[f"FQ{row}"].value = general_ws['H5'].value

        consolidated_ws[f"FR{row}"].value = general_ws['I4'].value
        consolidated_ws[f"FS{row}"].value = general_ws['I5'].value
        consolidated_ws[f"FT{row}"].value = general_ws['J4'].value
        consolidated_ws[f"FU{row}"].value = general_ws['J5'].value

        consolidated_ws[f"FV{row}"].value = general_ws['K4'].value
        consolidated_ws[f"FW{row}"].value = general_ws['K5'].value
        consolidated_ws[f"FX{row}"].value = general_ws['L4'].value
        consolidated_ws[f"FY{row}"].value = general_ws['L5'].value
        consolidated_ws[f"FZ{row}"].value = general_ws['M4'].value
        consolidated_ws[f"GA{row}"].value = general_ws['M5'].value
        consolidated_ws[f"GB{row}"].value = version
    
    
    consolidated_ws['GM2'].value = int(consolidated_ws['GM2'].value) + 1
    consolidated_ws['GN2'].value = int(consolidated_ws['GN2'].value) + 1
    consolidated_wb.save('Wireframe_Consolidated_Data.xlsx')
    return
    

def enclosure_extraction(path:str):
    """
        Purpose: 
            Extracts data from an Excel file in the Enclosure sheet and writes it to a consolidated Excel file.

        Parameters: 
            path (str): This is the file path of the Excel file.

        Returns:
            None: This function does not return anything, but it writes to the consolidated Excel file. 
    """
    excel_file_wb = load_workbook(path)
    enclosure_ws = excel_file_wb['1.Enclosure']

    consolidated_wb = load_workbook('Wireframe_Consolidated_Data.xlsx')
    consolidated_ws = consolidated_wb['Sheet']


    # get total number of parent rows in 1.Enclosure
    total_parent_rows = 0
    row = 8
    cell = enclosure_ws['B8'].value
    while cell is not None and str(cell).isdigit():
        total_parent_rows += 1
        row += 1
        cell = enclosure_ws[f"B{row}"].value


    # get asset ID
    asset_ID = retrieve_asset_ID(path)


    # recieving counter
    counter = consolidated_ws['GN2'].value


    # find the start of each table in 1.Enclosure
    max_rows = enclosure_ws.max_row
    start_of_table_list = []
    for row in range(1, max_rows + 1):
        if(enclosure_ws[f"B{row}"].value == 1):
            start_of_table_list.append(row)

    
    # filling in information from parent table
    row = consolidated_ws['GD2'].value
    row = int(row)
    total_new_rows = 0
    start_row = start_of_table_list[0]
    for row_num in range(start_row, start_row + total_parent_rows):
        row_list = []
        for column_letter in range(66, 85): 
            cell = enclosure_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)


        consolidated_ws[f"A{row}"].value = asset_ID
        consolidated_ws[f"B{row}"].value = '1.Enclosure'
        consolidated_ws[f"C{row}"].value = 'Parent'
        consolidated_ws[f"D{row}"].value = row_list[0]
        consolidated_ws[f"E{row}"].value = 0
        consolidated_ws[f"G{row}"].value = row_list[1]
        consolidated_ws[f"H{row}"].value = row_list[2]
        consolidated_ws[f"I{row}"].value = row_list[3]
        consolidated_ws[f"J{row}"].value = row_list[8]
        consolidated_ws[f"K{row}"].value = row_list[11]
        consolidated_ws[f"L{row}"].value = row_list[13]
        consolidated_ws[f"M{row}"].value = row_list[15]
        consolidated_ws[f"N{row}"].value = row_list[17]
        row += 1
        total_new_rows += 1


    # filling in information from second table
    row = consolidated_ws['GD2'].value
    row = int(row)
    start_row = start_of_table_list[1]
    for row_num in range(start_row, start_row + total_parent_rows):
        row_list = []
        for column_letter in range(66, 81): 
            # get first row from first table
            cell = enclosure_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        
        consolidated_ws[f"O{row}"].value = row_list[3]
        consolidated_ws[f"P{row}"].value = row_list[4]
        consolidated_ws[f"Q{row}"].value = row_list[5]
        consolidated_ws[f"R{row}"].value = row_list[6]
        consolidated_ws[f"S{row}"].value = row_list[7]
        consolidated_ws[f"T{row}"].value = row_list[9]
        consolidated_ws[f"U{row}"].value = row_list[10]
        consolidated_ws[f"V{row}"].value = row_list[11]
        consolidated_ws[f"W{row}"].value = counter
        row += 1


    # update data validation
    row = int(consolidated_ws['GD2'].value)
    for row_num in range(row, total_parent_rows + row):
        row_list = []
        for column_letter in range(71, 87):
            cell = consolidated_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        row_list.pop(1)
        if None not in row_list:
            consolidated_ws[f"F{row}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"F{row}"].value = 'Exclude'
        else:
            consolidated_ws[f"F{row}"].value = 'Incomplete'
        row += 1


    row = int(consolidated_ws['GD2'].value)
    consolidated_ws['GD2'].value = row + total_new_rows
    consolidated_wb.save('Wireframe_Consolidated_Data.xlsx')
    return 


def power_extraction(path:str):
    """
        Purpose:
            Extracts data from an Excel file in the Power sheet and writes it to a consolidated Excel file.

       Parameters: 
            path (str): This is the file path of the Excel file.

        Returns:
            None: This function does not return anything, but it writes to the consolidated Excel file. 
    """
    excel_file_wb = load_workbook(path)
    power_ws = excel_file_wb['2.Power']

    consolidated_wb = load_workbook('Wireframe_Consolidated_Data.xlsx')
    consolidated_ws = consolidated_wb['Sheet']


    # get asset ID
    asset_ID = retrieve_asset_ID(path)

    
    # recieving counter
    counter = consolidated_ws['GN2'].value


    # find the start of each table in 2.Power
    max_rows = power_ws.max_row
    start_of_table_list = []
    for row in range(1, max_rows + 1):
        if(power_ws[f"B{row}"].value == 1):
            start_of_table_list.append(row)


    # get total number of rows in parent table for 2.Power
    total_rows = 0
    pos = 8
    start_cell = power_ws['B8'].value
    while start_cell is not None and str(start_cell).isdigit():
        total_rows += 1
        pos += 1
        start_cell = power_ws[f"B{pos}"].value


    # filling in information for parent table
    row = consolidated_ws['GE2'].value
    row = int(row)
    start_row = start_of_table_list[0]
    total_new_parent_rows = 0
    for row_num in range(start_row, start_row + total_rows):
        row_list = []
        for column_letter in range(66, 78): 
            # get first row from first table
            cell = power_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        consolidated_ws[f"AA{row}"].value = asset_ID
        consolidated_ws[f"AB{row}"].value = '2.Power'
        consolidated_ws[f"AC{row}"].value = 'Parent'
        consolidated_ws[f"AD{row}"].value = row_list[0]
        consolidated_ws[f"AE{row}"].value = 0
        consolidated_ws[f"AG{row}"].value = row_list[1]
        consolidated_ws[f"AH{row}"].value = row_list[2]
        consolidated_ws[f"AI{row}"].value = row_list[6]
        consolidated_ws[f"AJ{row}"].value = row_list[8]
        consolidated_ws[f"AK{row}"].value = row_list[10]
        consolidated_ws[f"AL{row}"].value = row_list[11]
        consolidated_ws[f"AM{row}"].value = counter
        row += 1
        total_new_parent_rows += 1


    # filling in information for comms table
    row = int(consolidated_ws['GF2'].value)
    parent_row = 1
    child_row = 1
    total_new_child_rows = 0
    start_row = start_of_table_list[1]
    for row_num in range(start_row, max_rows):
        row_list = []
        for column_letter in range(66, 74): 
            cell = power_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        first_element = row_list[0]
        if first_element != None and first_element != 'Power Source for Comms and RTU' and first_element != 'Add Row Above':
            if first_element == 1 and row_num == start_row:
                child_row = 1
            elif first_element == 1:
                child_row = 1
                parent_row += 1
            
            
            consolidated_ws[f"AO{row}"].value = asset_ID
            consolidated_ws[f"AP{row}"].value = '2.Power'
            consolidated_ws[f"AQ{row}"].value = 'Comms'
            consolidated_ws[f"AR{row}"].value = parent_row
            consolidated_ws[f"AS{row}"].value = child_row
            consolidated_ws[f"AU{row}"].value = row_list[1]
            consolidated_ws[f"AV{row}"].value = row_list[2]
            consolidated_ws[f"AW{row}"].value = row_list[5]
            consolidated_ws[f"AX{row}"].value = row_list[7]
            consolidated_ws[f"AY{row}"].value = counter
            child_row += 1
            row += 1
            total_new_child_rows += 1


    # updating data validation for parent table
    start_row = consolidated_ws['GE2'].value
    start_row = int(start_row)
    for row_num in range(start_row, total_rows + start_row):
        row_list = []
        for column_letter in range(71, 77):
            cell = consolidated_ws['A' + chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        if None not in row_list:
            consolidated_ws[f"AF{row_num}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"AF{row_num}"].value = 'Exclude'
        else:
            consolidated_ws[f"AF{row_num}"].value = 'Incomplete'


    # updating data validation for comms table
    start_row = consolidated_ws['GF2'].value
    start_row = int(start_row)
    for row_num in range(start_row, total_new_child_rows + start_row):
        row_list = []
        for column_letter in range(85, 89):
            cell = consolidated_ws['A' + chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        if None not in row_list:
            consolidated_ws[f"AT{row_num}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"AT{row_num}"].value = 'Exclude'
        else:
            consolidated_ws[f"AT{row_num}"].value = 'Incomplete'
    
    
    row_b = consolidated_ws['GE2'].value
    row_c = consolidated_ws['GF2'].value
    row_b = int(row_b)
    row_c = int(row_c)
    consolidated_ws['GE2'].value = row_b + total_new_parent_rows
    consolidated_ws['GF2'].value = row_c + total_new_child_rows
    consolidated_wb.save('Wireframe_Consolidated_Data.xlsx')
    return


def instrumentation_extraction(path:str):
    """
        Purpose:
            Extracts data from an Excel file in the Instrumentation sheet and writes it to a consolidated Excel file.

        Parameters: 
            path (str): This is the file path of the Excel file.

        Returns:
            None: This function does not return anything, but it writes to the consolidated Excel file. 
    """
    excel_file_wb = load_workbook(path)
    instrumentation_ws = excel_file_wb['3.Instrumentation']

    consolidated_wb = load_workbook('Wireframe_Consolidated_Data.xlsx')
    consolidated_ws = consolidated_wb['Sheet']


    # get total number of rows in 3.Instrumentation
    total_rows = instrumentation_ws.max_row - 8


    # get asset ID
    asset_ID = retrieve_asset_ID(path)


    # recieving counter
    counter = consolidated_ws['GN2'].value


    # find the start of each table in 3.Instrumentation
    max_rows = instrumentation_ws.max_row
    start_of_table_list = []
    for row in range(1, max_rows + 1):
        if(instrumentation_ws[f"B{row}"].value == 1):
            start_of_table_list.append(row)


    # filling in information from first table
    row = int(consolidated_ws['GG2'].value)
    start_row = start_of_table_list[0]
    row_counter = 1
    total_new_rows = 0
    for row_num in range(start_row, start_row + total_rows):
        row_list = []
        for column_letter in range(66, 79): 
            cell = instrumentation_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)


        consolidated_ws[f"BA{row}"].value = asset_ID
        consolidated_ws[f"BB{row}"].value = '3.Instrumentation'
        consolidated_ws[f"BC{row}"].value = 'Parent'
        consolidated_ws[f"BD{row}"].value = row_counter
        consolidated_ws[f"BE{row}"].value = 0
        consolidated_ws[f"BG{row}"].value = row_list[1]
        consolidated_ws[f"BH{row}"].value = row_list[5]
        consolidated_ws[f"BI{row}"].value = row_list[9]
        consolidated_ws[f"BJ{row}"].value = row_list[11]
        consolidated_ws[f"BK{row}"].value = row_list[12]
        consolidated_ws[f"BL{row}"].value = counter

        row += 1
        row_counter += 1
        total_new_rows += 1


    # update data validation
    row = consolidated_ws['GG2'].value
    row = int(row)
    for row_num in range(row, total_rows + row):
        row_list = []
        for column_letter in range(71, 76):
            cell = consolidated_ws['B' + chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        if None not in row_list:
            consolidated_ws[f"BF{row}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"BF{row}"].value = 'Exclude'
        else:
            consolidated_ws[f"BF{row}"].value = 'Incomplete'
        
        row += 1

    consolidated_ws['GG2'].value = int(consolidated_ws['GG2'].value) + total_new_rows
    consolidated_wb.save('Wireframe_Consolidated_Data.xlsx')
    return


def plc_extraction(path:str):
    """
        Purpose:
            Extracts data from an Excel file in the PLC sheet and writes it to a consolidated Excel file.

        Parameters: 
            path (str): This is the file path of the Excel file.

        Returns:
            None: This function does not return anything, but it writes to the consolidated Excel file. 
    """
    excel_file_wb = load_workbook(path)
    plc_ws = excel_file_wb['4.PLC.RTU']

    consolidated_wb = load_workbook('Wireframe_Consolidated_Data.xlsx')
    consolidated_ws = consolidated_wb['Sheet']


    # get asset ID
    asset_ID = retrieve_asset_ID(path)


    
    # recieving counter
    counter = consolidated_ws['GN2'].value
    

    # find the start of each table in 4.PLC.RTU
    max_rows = plc_ws.max_row
    parent_table_list = []
    comms_table_list = []
    io_table_list = []
    for row in range(1, max_rows + 1):
        if(plc_ws[f"B{row}"].value == 1 and plc_ws[f"B{row-1}"].value == 'Port'):
            comms_table_list.append(row)
        elif(plc_ws[f"B{row}"].value == 1 and plc_ws[f"B{row-1}"].value == 'Slots'):
            io_table_list.append(row)
        elif(plc_ws[f"B{row}"].value == 1):
            parent_table_list.append(row)


    # inserting information for parent table
    start_row = parent_table_list[0]
    end_row = comms_table_list[0]
    parent_row = 1
    row = int(consolidated_ws['GH2'].value)
    total_parent_rows = 0 
    for row_number in range(start_row, end_row):
        row_list = []
        for column_letter in range(66, 79): 
            cell = plc_ws[chr(column_letter) + str(row_number)].value
            row_list.append(cell)
        
        first_element = str(row_list[0])
        if first_element.isdigit():
            consolidated_ws[f"CA{row}"].value = asset_ID
            consolidated_ws[f"CB{row}"].value = '4.PLC.RTU'
            consolidated_ws[f"CC{row}"].value = 'Parent'
            consolidated_ws[f"CD{row}"].value = parent_row
            consolidated_ws[f"CE{row}"].value = 0
            consolidated_ws[f"CG{row}"].value = row_list[1]
            consolidated_ws[f"CH{row}"].value = row_list[6]
            consolidated_ws[f"CI{row}"].value = row_list[10]
            consolidated_ws[f"CJ{row}"].value = row_list[12]
            consolidated_ws[f"CK{row}"].value = counter
            

            if row_list[1] == 'Other (specify)' and row_list[4] != None:
                consolidated_ws[f"CG{row}"].value = row_list[4]

            parent_row += 1
            row += 1
            total_parent_rows += 1


    # inserting information for comms table
    start_row = comms_table_list[0]
    end_row = io_table_list[0]
    parent_row = 1
    child_row = 1
    row = int(consolidated_ws['GI2'].value)
    total_comms_rows = 0
    for row_number in range(start_row, end_row - 1):
        row_list = []
        for column_letter in range(66, 75): 
            cell = plc_ws[chr(column_letter) + str(row_number)].value
            row_list.append(cell)

        first_element = str(row_list[0])

        if first_element != "None" and first_element != 'Port' and first_element != 'Add Row Above' and first_element != 'PLC/RTU Comms' and first_element != 'PLC/RTU - I/O':
            if first_element == '1' and row_number == start_row:
                child_row = 1
            elif first_element == '1':
                child_row = 1
                parent_row += 1

            consolidated_ws[f"CM{row}"].value = asset_ID
            consolidated_ws[f"CN{row}"].value = '4.PLC.RTU'
            consolidated_ws[f"CO{row}"].value = 'Comms'
            consolidated_ws[f"CP{row}"].value = parent_row
            consolidated_ws[f"CQ{row}"].value = child_row
            consolidated_ws[f"CS{row}"].value = row_list[1]
            consolidated_ws[f"CT{row}"].value = row_list[4]
            consolidated_ws[f"CU{row}"].value = row_list[6]
            consolidated_ws[f"CV{row}"].value = row_list[8]
            consolidated_ws[f"CW{row}"].value = counter
            child_row += 1
            row += 1
            total_comms_rows += 1


    # inserting information for IO table
    start_row = io_table_list[0]
    end_row = plc_ws.max_row
    parent_row = 1
    child_row = 1
    row = int(consolidated_ws['GJ2'].value)
    total_io_rows = 0
    for row_num in range(start_row, end_row):
        row_list = []
        for column_letter in range(66, 77): 
            cell = plc_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        first_element = row_list[0]
        if first_element != None and first_element != 'Slots' and first_element != 'Add Row Above' and first_element != 'PLC/RTU - I/O':
            if first_element == 1 and row_num == start_row:
                child_row = 1
            elif first_element == 1:
                child_row = 1
                parent_row += 1


            consolidated_ws[f"DA{row}"].value = asset_ID
            consolidated_ws[f"DB{row}"].value = '4.PLC.RTU'
            consolidated_ws[f"DC{row}"].value = 'IO'
            consolidated_ws[f"DD{row}"].value = parent_row
            consolidated_ws[f"DE{row}"].value = child_row
            consolidated_ws[f"DG{row}"].value = row_list[1]
            consolidated_ws[f"DH{row}"].value = row_list[3]
            consolidated_ws[f"DI{row}"].value = row_list[5]
            consolidated_ws[f"DJ{row}"].value = row_list[7]
            consolidated_ws[f"DK{row}"].value = row_list[10]
            consolidated_ws[f"DL{row}"].value = counter
            child_row += 1
            row += 1
            total_io_rows += 1


    # updating data validation for parent
    start_row = consolidated_ws['GH2'].value
    start_row = int(start_row)
    for row in range(start_row, total_parent_rows + start_row):
        row_list = []
        for column_letter in range(71, 75):
            cell = consolidated_ws['C' + chr(column_letter) + str(row)].value
            row_list.append(cell)

        if None not in row_list:
            consolidated_ws[f"CF{row}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"CF{row}"].value = 'Exclude'
        else:
            consolidated_ws[f"CF{row}"].value = 'Incomplete'


    # updating data validation for comms
    start_row = consolidated_ws['GI2'].value
    start_row = int(start_row)
    for row in range(start_row, total_comms_rows + start_row):
        row_list = []
        for column_letter in range(83, 87):
            cell = consolidated_ws['C' + chr(column_letter) + str(row)].value
            row_list.append(cell)

        if None not in row_list:
            consolidated_ws[f"CR{row}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"CR{row}"].value = 'Exclude'
        else:
            consolidated_ws[f"CR{row}"].value = 'Incomplete'


    # updating data validation for IO
    start_row = consolidated_ws['GJ2'].value
    start_row = int(start_row)
    for row in range(start_row, total_comms_rows + start_row):
        row_list = []
        for column_letter in range(71, 76):
            cell = consolidated_ws['D' + chr(column_letter) + str(row)].value
            row_list.append(cell)
        row_list.pop(1)

        if None not in row_list:
            consolidated_ws[f"DF{row}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"DF{row}"].value = 'Exclude'
        else:
            consolidated_ws[f"DF{row}"].value = 'Incomplete'

    consolidated_ws['GH2'].value = int(consolidated_ws['GH2'].value) + total_parent_rows
    consolidated_ws['GI2'].value = int(consolidated_ws['GI2'].value) + total_comms_rows
    consolidated_ws['GJ2'].value = int(consolidated_ws['GJ2'].value) + total_io_rows
    consolidated_wb.save('Wireframe_Consolidated_Data.xlsx')
    return 


def comms_extraction(path:str):
    """
        Purpose:
            Extracts data from an Excel file in the Comms sheet and writes it to a consolidated Excel file.

        Parameters: 
            path (str): This is the file path of the Excel file.

        Returns:
            None: This function does not return anything, but it writes to the consolidated Excel file. 
    """
    excel_file_wb = load_workbook(path)
    comms_ws = excel_file_wb['5.Comms']

    consolidated_wb = load_workbook('Wireframe_Consolidated_Data.xlsx')
    consolidated_ws = consolidated_wb['Sheet']


    # get asset ID
    asset_ID = retrieve_asset_ID(path)

    
    # recieving counter
    counter = consolidated_ws['GN2'].value


    # find the start of each table in 5.Comms
    max_rows = comms_ws.max_row
    start_of_table_list = []
    for row in range(1, max_rows + 1):
        if(comms_ws[f"B{row}"].value == 1):
            start_of_table_list.append(row)


    # get total number of rows in parent table for 2.Power
    total_parent_rows = 0
    pos = 8
    start_cell = comms_ws['B8'].value
    while start_cell is not None and str(start_cell).isdigit():
        total_parent_rows += 1
        pos += 1
        start_cell = comms_ws[f"B{pos}"].value


    # inserting information for parent table
    row = int(consolidated_ws['GK2'].value)
    start_row = start_of_table_list[0]
    for row_num in range(start_row, start_row + total_parent_rows):
        row_list = []
        for column_letter in range(66, 82): 
            cell = comms_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        
        consolidated_ws[f"EA{row}"].value = asset_ID
        consolidated_ws[f"EB{row}"].value = '5.Comms'
        consolidated_ws[f"EC{row}"].value = 'Parent'
        consolidated_ws[f"ED{row}"].value = row_list[0]
        consolidated_ws[f"EE{row}"].value = 0
        consolidated_ws[f"EG{row}"].value = row_list[1]
        consolidated_ws[f"EH{row}"].value = row_list[4]
        consolidated_ws[f"EI{row}"].value = row_list[6]
        consolidated_ws[f"EJ{row}"].value = row_list[10]
        consolidated_ws[f"EK{row}"].value = row_list[12]
        consolidated_ws[f"EL{row}"].value = row_list[14]
        consolidated_ws[f"EM{row}"].value = row_list[15]
        consolidated_ws[f"EN{row}"].value = counter
        row += 1


    # inserting information for comms tables
    row = consolidated_ws['GL2'].value
    row = int(row)
    parent_row = 1
    child_row = 1
    total_child_rows = 0
    start_row = start_of_table_list[1]
    for row_num in range(start_row, max_rows):
        row_list = []
        for column_letter in range(66, 75): 
            cell = comms_ws[chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        first_element = row_list[0]
        if first_element != None and first_element != 'Comm Ports' and first_element != 'Add Row Above' and first_element != 'Port':
            if first_element == 1 and row_num == start_row:
                child_row = 1
            elif first_element == 1:
                child_row = 1
                parent_row += 1
            
            
            consolidated_ws[f"EP{row}"].value = asset_ID
            consolidated_ws[f"EQ{row}"].value = '5.Comms'
            consolidated_ws[f"ER{row}"].value = 'Comms'
            consolidated_ws[f"ES{row}"].value = parent_row
            consolidated_ws[f"ET{row}"].value = child_row
            consolidated_ws[f"EV{row}"].value = row_list[1]
            consolidated_ws[f"EW{row}"].value = row_list[4]
            consolidated_ws[f"EX{row}"].value = row_list[6]
            consolidated_ws[f"EY{row}"].value = row_list[8]
            consolidated_ws[f"EZ{row}"].value = counter

            child_row += 1
            row += 1
            total_child_rows += 1


    # updating data validation for parent
    start_row = consolidated_ws['GK2'].value
    start_row = int(start_row)
    for row_num in range(start_row, total_parent_rows + start_row):
        row_list = []
        for column_letter in range(71, 78):
            cell = consolidated_ws['E' + chr(column_letter) + str(row_num)].value
            row_list.append(cell)

        row_list.pop(1)
        row_list.pop(2)

        if None not in row_list:
            consolidated_ws[f"EF{row_num}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"EF{row_num}"].value = 'Exclude'
        else:
            consolidated_ws[f"EF{row_num}"].value = 'Incomplete'


    # updating data validation for comms
    start_row = consolidated_ws['GL2'].value
    start_row = int(start_row)
    for row_num in range(start_row, total_child_rows + start_row):
        row_list = []
        for column_letter in range(86, 90):
            cell = consolidated_ws['E' + chr(column_letter) + str(row_num)].value
            row_list.append(cell)


        if None not in row_list:
            consolidated_ws[f"EU{row_num}"].value = 'Valid'
        elif all(item is None for item in row_list):
            consolidated_ws[f"EU{row_num}"].value = 'Exclude'
        else:
            consolidated_ws[f"EU{row_num}"].value = 'Incomplete'



    consolidated_ws['GK2'].value = int(consolidated_ws['GK2'].value) + total_parent_rows
    consolidated_ws['GL2'].value = int(consolidated_ws['GL2'].value) + total_child_rows
    consolidated_wb.save('Wireframe_Consolidated_Data.xlsx')
    return 


# a funciton that recieves the current time and date
def recieve_date_and_time():
    """
        Purpose: 
            Recieve the current date and time in military format.

        Parameters: 
            None.

        Returns:
            str: A string representing the current date and time in a specific format.
    """
    current_date = date.today()

    # Get the current time in the military format
    current_time = datetime.now().strftime("%H%M")

    return(str(current_date) + '_' + current_time + '_')


def create_duplicate_file(source_file_path:str, destination_file_path:str):
    """
        Purpose:
            Creates a duplicate of the Excel file from the source path to the destination path.

        Parameters:
            source_file_path (str): The path of the consolidated Excel file that will be duplicated.
            destination_file_path (str): The path of the duplicate file that will be created.

        Returns:
            None.
    """
    try:
        shutil.copyfile(source_file_path, destination_file_path)
    except FileNotFoundError:
        print(f"Source fild not found: {source_file_path}")
    except Exception as error:
        print(f"An error occurred: {error}")


def main_function(path_list:list):
    """
        Purpose:
            This function consolidated data from multiple Excel files and creates a new Excel file named 'Wirefram_Consolidated_Data.xlsx' containing the combined data. It also creates a duplicate copy of the consolidated file, appending the current date and time to the file name.

        Parameters:
            path_list (list): A list of file paths of the Excel files from which data needs to be consolidated.

        Returns:
            The function does not return anything explicity. It generates the consolidated data file and a duplicate copy with the current date and time in the file name.
    """
    # create a new excel file for the consolidated data
    if not os.path.exists("Wireframe_Consolidated_Data.xlsx"):
        create_excel_file("Wireframe_Consolidated_Data.xlsx")
    
    current_date_time = recieve_date_and_time()
    new_excel_file_name = f"{current_date_time}Wireframe_Consolidated_Data.xlsx"
    create_excel_file(new_excel_file_name)

    
    # will run through each of the excel files and insert the information into the consolidated file
    for path in path_list:
        general_extraction(path)
        enclosure_extraction(path)
        power_extraction(path)
        instrumentation_extraction(path)
        plc_extraction(path)
        comms_extraction(path)
        

    create_duplicate_file("Wireframe_Consolidated_Data.xlsx", new_excel_file_name)
        

# calling the functions
list = return_list_of_paths(folder_path)
main_function(list)
